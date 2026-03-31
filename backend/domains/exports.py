"""Exports domain API methods."""

import base64
import binascii
import os
import sqlite3
import subprocess
import sys
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Optional

from ..constants import (
    EXPORT_ALLOWED_FORMATS,
    EXPORT_ALLOWED_SECTIONS,
    EXPORT_SECTION_LABELS,
    config,
)
from ..helpers.api_response import JsonDict
from ..logger import logger
from .base import DomainApi


class ExportsApi(DomainApi):
    """Handles export generation for Excel/PDF reports."""

    def _normalize_bills_export_table(
        self,
        headers: list[str],
        rows: list[list[object]],
    ) -> tuple[list[str], list[list[object]]]:
        normalized_headers = [str(header).strip().lower() for header in headers]
        legacy_markers = {
            "id",
            "name",
            "amount",
            "paid_amount",
            "remaining_amount",
            "due_date",
            "status",
            "category_id",
            "category_name",
            "notes",
            "paid_at",
            "created_at",
            "payment_transaction_id",
        }

        if not any(marker in normalized_headers for marker in legacy_markers):
            return headers, rows

        index_by_header = {
            normalized: idx for idx, normalized in enumerate(normalized_headers)
        }

        def _pick_value(row: list[object], keys: tuple[str, ...]) -> object:
            for key in keys:
                idx = index_by_header.get(key)
                if idx is None:
                    continue
                if idx >= len(row):
                    continue
                return row[idx]
            return ""

        normalized_rows: list[list[object]] = []
        for row in rows:
            name = _pick_value(row, ("name", "nombre"))
            amount = _pick_value(row, ("amount", "monto"))
            paid_amount = _pick_value(row, ("paid_amount", "abonado"))
            remaining_amount = _pick_value(row, ("remaining_amount", "restante"))
            due_date = _pick_value(row, ("due_date", "vencimiento"))
            status = _pick_value(row, ("status", "estado"))
            category_name = _pick_value(
                row,
                ("category_name", "categoría", "categoria"),
            )
            paid_at = _pick_value(row, ("paid_at", "pagada en", "pagada_en"))
            notes = _pick_value(row, ("notes", "notas"))

            normalized_rows.append(
                [
                    name,
                    self._format_local_amount(amount),
                    self._format_local_amount(paid_amount),
                    self._format_local_amount(remaining_amount),
                    self._format_local_date(due_date),
                    self._label_bill_status(status),
                    category_name,
                    self._format_local_datetime(paid_at),
                    notes,
                ]
            )

        normalized_display_headers = [
            "Nombre",
            "Monto total",
            "Abonado",
            "Restante",
            "Vencimiento",
            "Estado",
            "Categoría",
            "Pagada en",
            "Notas",
        ]
        return normalized_display_headers, normalized_rows

    def _format_local_integer(self, value: object) -> str:
        try:
            integer = int(float(str(value)))
        except (TypeError, ValueError):
            return str(value)
        return f"{integer:,}".replace(",", ".")

    def _format_local_amount(self, value: object) -> str:
        try:
            amount = float(str(value))
        except (TypeError, ValueError):
            return str(value)

        formatted = f"{amount:,.2f}"
        formatted = formatted.replace(",", "_").replace(".", ",").replace("_", ".")
        return f"$ {formatted}"

    def _format_local_date(self, value: object) -> str:
        text = str(value).strip()
        if not text:
            return ""

        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
            try:
                return datetime.strptime(text, fmt).strftime("%d/%m/%Y")
            except ValueError:
                continue

        return text

    def _format_local_datetime(self, value: object) -> str:
        text = str(value).strip()
        if not text:
            return ""

        for fmt in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%dT%H:%M:%S",
            "%Y-%m-%dT%H:%M",
        ):
            try:
                return datetime.strptime(text, fmt).strftime("%d/%m/%Y %H:%M")
            except ValueError:
                continue

        if len(text) >= 10:
            try:
                parsed_date = datetime.strptime(text[:10], "%Y-%m-%d")
                return parsed_date.strftime("%d/%m/%Y")
            except ValueError:
                pass

        return text

    def _label_transaction_type(self, value: object) -> str:
        normalized = str(value).strip().lower()
        if normalized == "income":
            return "Ingreso"
        if normalized == "expense":
            return "Gasto"
        return str(value)

    def _label_category_type(self, value: object) -> str:
        normalized = str(value).strip().lower()
        if normalized == "income":
            return "Ingreso"
        if normalized == "expense":
            return "Gasto"
        return str(value)

    def _label_bill_status(self, value: object) -> str:
        normalized = str(value).strip().lower()
        if normalized == "pending":
            return "Pendiente"
        if normalized == "paid":
            return "Pagada"
        if normalized == "overdue":
            return "Vencida"
        return str(value)

    def _validate_user(self, user_id: int) -> JsonDict | None:
        if user_id <= 0:
            return self._error("El user_id debe ser mayor a 0")

        try:
            cur = self.conn.cursor()
            cur.execute("SELECT id FROM users WHERE id = ?", (user_id,))
            if cur.fetchone() is None:
                return self._error("Usuario no encontrado")
            return None
        except sqlite3.Error as exc:
            logger.error("Error validating user for exports: %s", exc)
            return self._error("Error interno al validar usuario")

    def _normalize_format(self, export_format: str) -> str:
        return export_format.strip().lower()

    def _normalize_section(self, section: str) -> str:
        return section.strip().lower()

    def _validate_format_and_section(self, export_format: str, section: str) -> JsonDict | None:
        normalized_format = self._normalize_format(export_format)
        normalized_section = self._normalize_section(section)

        if normalized_format not in EXPORT_ALLOWED_FORMATS:
            supported_formats = ", ".join(sorted(EXPORT_ALLOWED_FORMATS))
            return self._error(
                f"Formato inválido. Usa uno de: {supported_formats}"
            )

        if normalized_section not in EXPORT_ALLOWED_SECTIONS:
            supported_sections = ", ".join(sorted(EXPORT_ALLOWED_SECTIONS))
            return self._error(
                f"Sección inválida. Usa una de: {supported_sections}"
            )

        return None

    def _get_export_dir(self, user_id: int) -> Path:
        export_dir = config.exports_dir / f"user_{user_id}"
        export_dir.mkdir(parents=True, exist_ok=True)
        return export_dir

    def _decode_png_data_url(self, image_data_url: str) -> bytes:
        if not isinstance(image_data_url, str) or len(image_data_url.strip()) == 0:
            raise ValueError("No se recibió una imagen válida para exportar")

        prefix = "data:image/png;base64,"
        if not image_data_url.startswith(prefix):
            raise ValueError("Formato de imagen inválido. Se esperaba PNG en base64")

        encoded = image_data_url[len(prefix):]
        try:
            image_bytes = base64.b64decode(encoded, validate=True)
        except (ValueError, binascii.Error) as exc:
            raise ValueError("No se pudo decodificar la imagen del gráfico") from exc

        if len(image_bytes) == 0:
            raise ValueError("La imagen del gráfico está vacía")

        return image_bytes

    def _fetch_transactions(
        self,
        user_id: int,
        year: Optional[int] = None,
        month: Optional[int] = None,
        from_date: Optional[str] = None,
    ) -> list[JsonDict]:
        cur = self.conn.cursor()
        if from_date is not None:
            cur.execute(
                """
                SELECT
                    t.id,
                    t.date,
                    t.type,
                    t.amount,
                    t.description,
                    t.credit_id,
                    c.name AS category_name
                FROM transactions t
                LEFT JOIN categories c ON c.id = t.category_id
                WHERE t.user_id = ?
                  AND date(t.date) >= date(?)
                ORDER BY t.date DESC, t.id DESC
                """,
                (user_id, from_date),
            )
        elif year is not None and month is not None:
            cur.execute(
                """
                SELECT
                    t.id,
                    t.date,
                    t.type,
                    t.amount,
                    t.description,
                    t.credit_id,
                    c.name AS category_name
                FROM transactions t
                LEFT JOIN categories c ON c.id = t.category_id
                WHERE t.user_id = ?
                  AND strftime('%Y', t.date) = ?
                  AND strftime('%m', t.date) = ?
                ORDER BY t.date DESC, t.id DESC
                """,
                (user_id, str(year), f"{month:02d}"),
            )
        else:
            cur.execute(
                """
                SELECT
                    t.id,
                    t.date,
                    t.type,
                    t.amount,
                    t.description,
                    t.credit_id,
                    c.name AS category_name
                FROM transactions t
                LEFT JOIN categories c ON c.id = t.category_id
                WHERE t.user_id = ?
                ORDER BY t.date DESC, t.id DESC
                """,
                (user_id,),
            )

        return [
            {
                "id": int(row["id"]),
                "date": row["date"],
                "type": row["type"],
                "amount": float(row["amount"]),
                "category_name": row["category_name"] or "Sin categoría",
                "description": row["description"] or "",
                "credit_id": row["credit_id"],
            }
            for row in cur.fetchall()
        ]

    def _fetch_credits(self, user_id: int) -> list[JsonDict]:
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT
                c.id,
                c.description,
                c.total_amount,
                c.installments,
                c.installment_amount,
                c.paid_installments,
                c.start_date,
                cat.name AS category_name
            FROM credits c
            LEFT JOIN categories cat ON cat.id = c.category_id
            WHERE c.user_id = ?
            ORDER BY c.created_at DESC, c.id DESC
            """,
            (user_id,),
        )

        rows: list[JsonDict] = []
        for row in cur.fetchall():
            installments = int(row["installments"])
            paid_installments = int(row["paid_installments"])
            rows.append(
                {
                    "id": int(row["id"]),
                    "description": row["description"],
                    "total_amount": float(row["total_amount"]),
                    "installments": installments,
                    "installment_amount": float(row["installment_amount"]),
                    "paid_installments": paid_installments,
                    "pending_installments": max(installments - paid_installments, 0),
                    "start_date": row["start_date"],
                    "category_name": row["category_name"] or "Sin categoría",
                }
            )

        return rows

    def _fetch_categories(self, user_id: int) -> list[JsonDict]:
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT
                c.id,
                c.name,
                c.type,
                c.color,
                COUNT(t.id) AS transactions_count,
                COALESCE(SUM(t.amount), 0) AS total_amount
            FROM categories c
            LEFT JOIN transactions t
              ON t.category_id = c.id
             AND t.user_id = c.user_id
            WHERE c.user_id = ?
            GROUP BY c.id, c.name, c.type, c.color
            ORDER BY c.type ASC, c.name ASC
            """,
            (user_id,),
        )

        return [
            {
                "id": int(row["id"]),
                "name": row["name"],
                "type": row["type"],
                "color": row["color"],
                "transactions_count": int(row["transactions_count"]),
                "total_amount": float(row["total_amount"]),
            }
            for row in cur.fetchall()
        ]

    def _fetch_notes(self, user_id: int) -> list[JsonDict]:
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT id, title, content, created_at, updated_at
            FROM notes
            WHERE user_id = ?
            ORDER BY datetime(updated_at) DESC, id DESC
            """,
            (user_id,),
        )

        return [
            {
                "id": int(row["id"]),
                "title": row["title"],
                "content": row["content"] or "",
                "created_at": row["created_at"],
                "updated_at": row["updated_at"],
            }
            for row in cur.fetchall()
        ]

    def _fetch_bills(self, user_id: int) -> list[JsonDict]:
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT
                b.id,
                b.name,
                b.amount,
                COALESCE(b.paid_amount, 0) AS paid_amount,
                (b.amount - COALESCE(b.paid_amount, 0)) AS remaining_amount,
                b.due_date,
                b.status,
                c.name AS category_name,
                b.notes,
                b.paid_at,
                b.created_at
            FROM bills b
            LEFT JOIN categories c ON c.id = b.category_id
            WHERE b.user_id = ?
            ORDER BY date(b.due_date) ASC, b.id ASC
            """,
            (user_id,),
        )

        return [
            {
                "id": int(row["id"]),
                "name": row["name"],
                "amount": float(row["amount"]),
                "paid_amount": float(row["paid_amount"]),
                "remaining_amount": max(float(row["remaining_amount"]), 0.0),
                "due_date": row["due_date"],
                "status": row["status"],
                "category_name": row["category_name"] or "Sin categoría",
                "notes": row["notes"] or "",
                "paid_at": row["paid_at"] or "",
                "created_at": row["created_at"],
            }
            for row in cur.fetchall()
        ]

    def _fetch_summary(self, user_id: int) -> list[JsonDict]:
        cur = self.conn.cursor()

        cur.execute(
            """
            SELECT
                COUNT(*) AS transactions_count,
                COALESCE(SUM(CASE WHEN type = 'income' THEN amount ELSE 0 END), 0) AS income_total,
                COALESCE(SUM(CASE WHEN type = 'expense' THEN amount ELSE 0 END), 0) AS expense_total
            FROM transactions
            WHERE user_id = ?
            """,
            (user_id,),
        )
        tx_row = cur.fetchone()

        cur.execute(
            "SELECT COUNT(*) AS credits_count FROM credits WHERE user_id = ?",
            (user_id,),
        )
        credits_row = cur.fetchone()

        cur.execute(
            "SELECT COUNT(*) AS categories_count FROM categories WHERE user_id = ?",
            (user_id,),
        )
        categories_row = cur.fetchone()

        cur.execute(
            "SELECT COUNT(*) AS notes_count FROM notes WHERE user_id = ?",
            (user_id,),
        )
        notes_row = cur.fetchone()

        cur.execute(
            """
            SELECT
                COUNT(*) AS bills_count,
                COALESCE(SUM(CASE WHEN status = 'pending' THEN 1 ELSE 0 END), 0) AS pending_bills_count,
                COALESCE(SUM(CASE WHEN status = 'overdue' THEN 1 ELSE 0 END), 0) AS overdue_bills_count,
                COALESCE(SUM(CASE WHEN status = 'paid' THEN 1 ELSE 0 END), 0) AS paid_bills_count,
                COALESCE(
                    SUM(
                        CASE
                            WHEN status IN ('pending', 'overdue')
                            THEN (amount - COALESCE(paid_amount, 0))
                            ELSE 0
                        END
                    ),
                    0
                ) AS open_bills_amount
            FROM bills
            WHERE user_id = ?
            """,
            (user_id,),
        )
        bills_row = cur.fetchone()

        cur.execute(
            """
            SELECT
                COALESCE(c.name, 'Sin categoría') AS category_name,
                COALESCE(SUM(t.amount), 0) AS total
            FROM transactions t
            LEFT JOIN categories c ON c.id = t.category_id
            WHERE t.user_id = ?
              AND t.type = 'expense'
            GROUP BY t.category_id
            ORDER BY total DESC
            LIMIT 1
            """,
            (user_id,),
        )
        top_expense = cur.fetchone()

        income_total = float(tx_row["income_total"])
        expense_total = float(tx_row["expense_total"])
        balance = income_total - expense_total
        pending_bills_count = int(bills_row["pending_bills_count"])
        overdue_bills_count = int(bills_row["overdue_bills_count"])
        open_bills_count = pending_bills_count + overdue_bills_count
        open_bills_amount = max(float(bills_row["open_bills_amount"]), 0.0)

        return [
            {"metric": "Fecha de generación", "value": datetime.now().strftime("%d/%m/%Y %H:%M:%S")},
            {"metric": "Total de transacciones", "value": int(tx_row["transactions_count"])},
            {"metric": "Ingresos acumulados", "value": round(income_total, 2)},
            {"metric": "Gastos acumulados", "value": round(expense_total, 2)},
            {"metric": "Balance", "value": round(balance, 2)},
            {"metric": "Créditos", "value": int(credits_row["credits_count"])},
            {"metric": "Categorías", "value": int(categories_row["categories_count"])},
            {"metric": "Notas", "value": int(notes_row["notes_count"])},
            {"metric": "Facturas", "value": int(bills_row["bills_count"])} ,
            {"metric": "Facturas abiertas", "value": open_bills_count},
            {"metric": "Facturas vencidas", "value": overdue_bills_count},
            {"metric": "Facturas pagadas", "value": int(bills_row["paid_bills_count"])},
            {"metric": "Saldo facturas pendientes", "value": round(open_bills_amount, 2)},
            {
                "metric": "Mayor categoría de gasto",
                "value": (
                    f"{top_expense['category_name']} ({self._format_local_amount(top_expense['total'])})"
                    if top_expense is not None
                    else "Sin datos"
                ),
            },
        ]

    def _build_headers_and_rows(
        self,
        section: str,
        user_id: int,
        year: Optional[int] = None,
        month: Optional[int] = None,
        from_date: Optional[str] = None,
    ) -> tuple[list[str], list[list[object]]]:
        if section == "transactions":
            transactions = self._fetch_transactions(
                user_id,
                year=year,
                month=month,
                from_date=from_date,
            )
            headers = [
                "ID",
                "Fecha",
                "Tipo",
                "Monto",
                "Categoría",
                "Descripción",
                "ID Crédito",
            ]
            rows = [
                [
                    tx["id"],
                    self._format_local_date(tx["date"]),
                    self._label_transaction_type(tx["type"]),
                    self._format_local_amount(tx["amount"]),
                    tx["category_name"],
                    tx["description"],
                    tx["credit_id"] if tx["credit_id"] is not None else "",
                ]
                for tx in transactions
            ]
            return headers, rows

        if section == "credits":
            credits = self._fetch_credits(user_id)
            headers = [
                "ID",
                "Descripción",
                "Monto Total",
                "Cuotas",
                "Monto por Cuota",
                "Cuotas Pagadas",
                "Cuotas Pendientes",
                "Fecha Inicio",
                "Categoría",
            ]
            rows = [
                [
                    credit["id"],
                    credit["description"],
                    self._format_local_amount(credit["total_amount"]),
                    self._format_local_integer(credit["installments"]),
                    self._format_local_amount(credit["installment_amount"]),
                    self._format_local_integer(credit["paid_installments"]),
                    self._format_local_integer(credit["pending_installments"]),
                    self._format_local_date(credit["start_date"]),
                    credit["category_name"],
                ]
                for credit in credits
            ]
            return headers, rows

        if section == "categories":
            categories = self._fetch_categories(user_id)
            headers = [
                "ID",
                "Nombre",
                "Tipo",
                "Color",
                "Cantidad de Transacciones",
                "Monto Total",
            ]
            rows = [
                [
                    category["id"],
                    category["name"],
                    self._label_category_type(category["type"]),
                    category["color"],
                    self._format_local_integer(category["transactions_count"]),
                    self._format_local_amount(category["total_amount"]),
                ]
                for category in categories
            ]
            return headers, rows

        if section == "notes":
            notes = self._fetch_notes(user_id)
            headers = ["ID", "Título", "Contenido", "Creada", "Actualizada"]
            rows = [
                [
                    note["id"],
                    note["title"],
                    note["content"],
                    self._format_local_datetime(note["created_at"]),
                    self._format_local_datetime(note["updated_at"]),
                ]
                for note in notes
            ]
            return headers, rows

        if section == "bills":
            bills = self._fetch_bills(user_id)
            headers = [
                "Nombre",
                "Monto total",
                "Abonado",
                "Restante",
                "Vencimiento",
                "Estado",
                "Categoría",
                "Pagada en",
                "Notas",
            ]
            rows = [
                [
                    bill["name"],
                    self._format_local_amount(bill["amount"]),
                    self._format_local_amount(bill["paid_amount"]),
                    self._format_local_amount(bill["remaining_amount"]),
                    self._format_local_date(bill["due_date"]),
                    self._label_bill_status(bill["status"]),
                    bill["category_name"],
                    self._format_local_datetime(bill["paid_at"]),
                    bill["notes"],
                ]
                for bill in bills
            ]
            return headers, rows

        summary = self._fetch_summary(user_id)
        headers = ["Métrica", "Valor"]
        rows: list[list[object]] = []
        for item in summary:
            metric = str(item["metric"])
            value = item["value"]

            if metric in {
                "Total de transacciones",
                "Créditos",
                "Categorías",
                "Notas",
                "Facturas",
                "Facturas abiertas",
                "Facturas vencidas",
                "Facturas pagadas",
            }:
                value = self._format_local_integer(value)
            elif metric in {
                "Ingresos acumulados",
                "Gastos acumulados",
                "Balance",
                "Saldo facturas pendientes",
            }:
                value = self._format_local_amount(value)
            elif metric == "Fecha de generación":
                value = self._format_local_datetime(value)

            rows.append([metric, value])
        return headers, rows

    def _write_excel(self, file_path: Path, headers: list[str], rows: list[list[object]]) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ModuleNotFoundError as exc:
            raise RuntimeError(
                "No se pudo generar Excel porque falta la dependencia 'openpyxl'"
            ) from exc

        workbook = Workbook()
        sheet = workbook.active
        if sheet is None:
            raise RuntimeError("No se pudo crear la hoja del archivo Excel")
        sheet.title = "Exportacion"

        sheet.append(headers)
        for row in rows:
            sheet.append(["" if value is None else value for value in row])

        header_fill = PatternFill(fill_type="solid", start_color="3E2C8E", end_color="3E2C8E")
        header_font = Font(color="FFFFFF", bold=True)

        for col_index, header in enumerate(headers, start=1):
            header_cell = sheet.cell(row=1, column=col_index)
            header_cell.fill = header_fill
            header_cell.font = header_font
            header_cell.alignment = Alignment(horizontal="center", vertical="center")

            max_len = len(str(header))
            for row_index in range(2, len(rows) + 2):
                value = sheet.cell(row=row_index, column=col_index).value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))

            col_letter = get_column_letter(col_index)
            sheet.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)

        sheet.freeze_panes = "A2"
        workbook.save(file_path)

    def _get_pdf_column_widths(
        self,
        section: str,
        headers_count: int,
        available_width: float,
    ) -> list[float]:
        ratios_by_section = {
            "summary": [0.35, 0.65],
            "transactions": [0.07, 0.12, 0.09, 0.1, 0.16, 0.36, 0.1],
            "credits": [0.05, 0.22, 0.1, 0.08, 0.1, 0.1, 0.1, 0.11, 0.14],
            "bills": [0.14, 0.11, 0.11, 0.11, 0.1, 0.1, 0.12, 0.09, 0.12],
            "categories": [0.08, 0.27, 0.12, 0.12, 0.19, 0.22],
            "notes": [0.06, 0.24, 0.46, 0.12, 0.12],
        }
        ratios = ratios_by_section.get(section, [])

        if len(ratios) != headers_count:
            return [available_width / max(headers_count, 1)] * headers_count

        return [available_width * ratio for ratio in ratios]

    def _write_pdf(
        self,
        file_path: Path,
        section: str,
        headers: list[str],
        rows: list[list[object]],
    ) -> None:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        except ModuleNotFoundError as exc:
            raise RuntimeError(
                "No se pudo generar PDF porque falta la dependencia 'reportlab'"
            ) from exc

        section_name = EXPORT_SECTION_LABELS.get(section, section)
        is_wide_section = section in {"transactions", "credits", "notes", "bills"}
        page_size = landscape(A4) if is_wide_section else A4

        left_margin = right_margin = 12 * mm
        top_margin = 14 * mm
        bottom_margin = 14 * mm

        doc = SimpleDocTemplate(
            str(file_path),
            pagesize=page_size,
            leftMargin=left_margin,
            rightMargin=right_margin,
            topMargin=top_margin,
            bottomMargin=bottom_margin,
            title=f"Bolsi - {section_name}",
            author="Bolsi",
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "ExportTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=16,
            leading=20,
            spaceAfter=4,
            alignment=0,
        )
        meta_style = ParagraphStyle(
            "ExportMeta",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#5C5C5C"),
        )
        table_header_style = ParagraphStyle(
            "TableHeader",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            textColor=colors.white,
        )
        table_cell_style = ParagraphStyle(
            "TableCell",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
        )

        available_width = doc.width
        col_widths = self._get_pdf_column_widths(section, len(headers), available_width)

        def _paragraph_text(value: object, style: ParagraphStyle) -> Paragraph:
            text = str(value) if value is not None else ""
            escaped = (
                text.replace("&", "&amp;")
                .replace("<", "&lt;")
                .replace(">", "&gt;")
                .replace("\n", "<br/>")
            )
            return Paragraph(escaped, style)

        table_data: list[list[Paragraph]] = [
            [_paragraph_text(header, table_header_style) for header in headers]
        ]
        table_data.extend(
            [[_paragraph_text(item, table_cell_style) for item in row] for row in rows]
        )

        table = Table(
            table_data,
            colWidths=col_widths,
            repeatRows=1,
            splitByRow=1,
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3E2C8E")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 8),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F7F6FF")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#FFFFFF"), colors.HexColor("#F6F5FF")]),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#C5B9F5")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )

        story = [
            Paragraph(f"Bolsi - Exportación {section_name}", title_style),
            Paragraph(
                f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} · Registros: {len(rows)}",
                meta_style,
            ),
            Spacer(1, 8),
            table,
        ]
        doc.build(story)

    def _write_dashboard_visual_pdf(
        self,
        file_path: Path,
        image_data_url: str,
        period_label: str,
        generated_at: str,
        month_income: float,
        month_expense: float,
        month_balance: float,
        active_credits: int,
        pending_installments: int,
        monthly_due_amount: float,
        bills_count: int,
        overdue_bills_count: int,
        due_soon_bills_count: int,
        month_bills_amount: float,
        categories_count: int,
    ) -> None:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.lib.utils import ImageReader
            from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        except ModuleNotFoundError as exc:
            raise RuntimeError(
                "No se pudo generar PDF porque falta la dependencia 'reportlab'"
            ) from exc

        image_bytes = self._decode_png_data_url(image_data_url)

        doc = SimpleDocTemplate(
            str(file_path),
            pagesize=A4,
            leftMargin=14 * mm,
            rightMargin=14 * mm,
            topMargin=14 * mm,
            bottomMargin=14 * mm,
            title="Bolsi - Dashboard visual",
            author="Bolsi",
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "DashboardVisualTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=16,
            leading=20,
            spaceAfter=4,
        )
        meta_style = ParagraphStyle(
            "DashboardVisualMeta",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#5C5C5C"),
        )
        section_style = ParagraphStyle(
            "DashboardVisualSection",
            parent=styles["Heading3"],
            fontName="Helvetica-Bold",
            fontSize=11,
            leading=14,
            spaceBefore=4,
            spaceAfter=4,
        )

        metrics_table = Table(
            [
                ["Métrica", "Valor"],
                ["Ingresos del mes", f"{month_income:.2f}"],
                ["Gastos del mes", f"{month_expense:.2f}"],
                ["Balance del mes", f"{month_balance:.2f}"],
                ["Créditos activos", str(active_credits)],
                ["Cuotas pendientes", str(pending_installments)],
                ["Cuotas a pagar este mes", f"{monthly_due_amount:.2f}"],
                ["Facturas del mes", str(bills_count)],
                ["Facturas vencidas", str(overdue_bills_count)],
                ["Facturas por vencer", str(due_soon_bills_count)],
                ["Monto facturas del mes", f"{month_bills_amount:.2f}"],
                ["Categorías creadas", str(categories_count)],
            ],
            colWidths=[doc.width * 0.45, doc.width * 0.55],
            repeatRows=1,
        )
        metrics_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3E2C8E")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F7F6FF")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#FFFFFF"), colors.HexColor("#F6F5FF")]),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#C5B9F5")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )

        image_buffer = BytesIO(image_bytes)
        image_reader = ImageReader(image_buffer)
        image_width, image_height = image_reader.getSize()

        max_width = doc.width
        max_height = 95 * mm

        scaled_width = max_width
        scaled_height = scaled_width * (image_height / max(image_width, 1))

        if scaled_height > max_height:
            scaled_height = max_height
            scaled_width = scaled_height * (image_width / max(image_height, 1))

        image_buffer.seek(0)
        chart_image = Image(image_buffer, width=scaled_width, height=scaled_height)
        chart_image.hAlign = "CENTER"

        story = [
            Paragraph("Bolsi - Dashboard visual", title_style),
            Paragraph(f"Periodo: {period_label}", meta_style),
            Paragraph(f"Generado: {generated_at}", meta_style),
            Spacer(1, 8),
            Paragraph("Metricas del mes", section_style),
            metrics_table,
            Spacer(1, 10),
            Paragraph("Graficos del dashboard", section_style),
            chart_image,
        ]
        doc.build(story)

    def export_dashboard_chart_png(self, user_id: int, image_data_url: str) -> JsonDict:
        user_error = self._validate_user(user_id)
        if user_error:
            return user_error

        try:
            image_bytes = self._decode_png_data_url(image_data_url)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_name = f"bolsi_dashboard_chart_{timestamp}.png"
            export_dir = self._get_export_dir(user_id)
            file_path = export_dir / file_name
            file_path.write_bytes(image_bytes)

            return self._success(
                "Exportación generada correctamente",
                data={
                    "file_path": str(file_path),
                    "file_name": file_name,
                    "format": "png",
                    "section": "dashboard_chart",
                },
            )
        except ValueError as exc:
            logger.error("Invalid chart image for PNG export: %s", exc)
            return self._error(str(exc))
        except OSError as exc:
            logger.error("Error writing dashboard PNG export: %s", exc)
            return self._error("Error al guardar el archivo exportado")
        except Exception as exc:
            logger.error("Unexpected error exporting dashboard PNG: %s", exc)
            return self._error("Error interno al generar exportación")

    def export_dashboard_visual_pdf(
        self,
        user_id: int,
        image_data_url: str,
        period_label: str,
        generated_at: str,
        month_income: float,
        month_expense: float,
        month_balance: float,
        active_credits: int,
        pending_installments: int,
        monthly_due_amount: float,
        bills_count: int,
        overdue_bills_count: int,
        due_soon_bills_count: int,
        month_bills_amount: float,
        categories_count: int,
    ) -> JsonDict:
        user_error = self._validate_user(user_id)
        if user_error:
            return user_error

        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_name = f"bolsi_dashboard_visual_{timestamp}.pdf"
            export_dir = self._get_export_dir(user_id)
            file_path = export_dir / file_name

            self._write_dashboard_visual_pdf(
                file_path,
                image_data_url,
                period_label,
                generated_at,
                float(month_income),
                float(month_expense),
                float(month_balance),
                int(active_credits),
                int(pending_installments),
                float(monthly_due_amount),
                int(bills_count),
                int(overdue_bills_count),
                int(due_soon_bills_count),
                float(month_bills_amount),
                int(categories_count),
            )

            return self._success(
                "Exportación generada correctamente",
                data={
                    "file_path": str(file_path),
                    "file_name": file_name,
                    "format": "pdf",
                    "section": "dashboard_visual",
                },
            )
        except ValueError as exc:
            logger.error("Invalid dashboard visual data: %s", exc)
            return self._error(str(exc))
        except OSError as exc:
            logger.error("Error writing dashboard visual PDF: %s", exc)
            return self._error("Error al guardar el archivo exportado")
        except RuntimeError as exc:
            logger.error("Missing dependency for dashboard visual export: %s", exc)
            return self._error(str(exc))
        except Exception as exc:
            logger.error("Unexpected error exporting dashboard visual PDF: %s", exc)
            return self._error("Error interno al generar exportación")

    def open_export_folder(self, user_id: int) -> JsonDict:
        user_error = self._validate_user(user_id)
        if user_error:
            return user_error

        export_dir = self._get_export_dir(user_id).resolve()
        export_dir_str = str(export_dir)

        try:
            if sys.platform.startswith("win"):
                os.startfile(export_dir_str)  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                subprocess.Popen(["open", export_dir_str])
            else:
                subprocess.Popen(["xdg-open", export_dir_str])

            return self._success(
                "Carpeta de exportaciones abierta",
                data={"folder_path": export_dir_str},
            )
        except OSError as exc:
            logger.error("Error opening exports folder '%s': %s", export_dir, exc)
            return self._error("No se pudo abrir la carpeta de exportaciones")
        except Exception as exc:
            logger.error("Unexpected error opening exports folder '%s': %s", export_dir, exc)
            return self._error("Error interno al abrir la carpeta de exportaciones")

    def generate_export(
        self,
        user_id: int,
        section: str = "summary",
        export_format: str = "xlsx",
        year: Optional[int] = None,
        month: Optional[int] = None,
        from_date: Optional[str] = None,
    ) -> JsonDict:
        user_error = self._validate_user(user_id)
        if user_error:
            return user_error

        normalized_section = self._normalize_section(section)
        normalized_format = self._normalize_format(export_format)

        validation_error = self._validate_format_and_section(
            normalized_format,
            normalized_section,
        )
        if validation_error:
            return validation_error

        if normalized_section == "transactions":
            if from_date is not None:
                if year is not None or month is not None:
                    return self._error("Usa from_date o year/month para transacciones, no ambos")

                normalized_from_date = from_date.strip()
                if not normalized_from_date:
                    return self._error("from_date no puede estar vacío")

                try:
                    date.fromisoformat(normalized_from_date)
                except ValueError:
                    return self._error("from_date debe tener formato YYYY-MM-DD")

                from_date = normalized_from_date

            has_year = year is not None
            has_month = month is not None
            if has_year != has_month:
                return self._error("Para exportar transacciones filtradas debes indicar year y month")

            if has_year and has_month:
                if year < 1900 or year > 9999:
                    return self._error("El year es inválido")
                if month < 1 or month > 12:
                    return self._error("El month debe estar entre 1 y 12")

        try:
            headers, rows = self._build_headers_and_rows(
                normalized_section,
                user_id,
                year=year,
                month=month,
                from_date=from_date,
            )
            if normalized_section == "bills":
                headers, rows = self._normalize_bills_export_table(headers, rows)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            extension = normalized_format
            period_suffix = ""
            if normalized_section == "transactions":
                if from_date is not None:
                    period_suffix = f"_from_{from_date.replace('-', '')}"
                elif year is not None and month is not None:
                    period_suffix = f"_{year}_{month:02d}"
            file_name = f"bolsi_{normalized_section}{period_suffix}_{timestamp}.{extension}"
            export_dir = self._get_export_dir(user_id)
            file_path = export_dir / file_name

            if normalized_format == "xlsx":
                self._write_excel(file_path, headers, rows)
            else:
                self._write_pdf(file_path, normalized_section, headers, rows)

            return self._success(
                "Exportación generada correctamente",
                data={
                    "file_path": str(file_path),
                    "file_name": file_name,
                    "format": normalized_format,
                    "section": normalized_section,
                    "row_count": len(rows),
                },
            )
        except sqlite3.Error as exc:
            logger.error("Error querying data for export: %s", exc)
            return self._error("Error interno al preparar la exportación")
        except OSError as exc:
            logger.error("Error writing export file: %s", exc)
            return self._error("Error al guardar el archivo exportado")
        except RuntimeError as exc:
            logger.error("Missing dependency for export: %s", exc)
            return self._error(str(exc))
        except Exception as exc:
            logger.error("Unexpected error generating export: %s", exc)
            return self._error("Error interno al generar exportación")
